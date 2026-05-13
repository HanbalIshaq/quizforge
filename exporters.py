import csv
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def attempts_to_csv(attempts: list[dict], questions: list[dict]) -> str:
    buf = io.StringIO()
    w = csv.writer(buf)
    header = ["Name", "Email", "Score", "Max", "Percentage", "Started", "Submitted"]
    for q in questions:
        header.append(f"Q{q['position']+1}: {q['text'][:60]}")
    w.writerow(header)
    for a in attempts:
        row = [
            a.get("student_name", ""),
            a.get("student_email", ""),
            a.get("score", 0),
            a.get("max_score", 0),
            f"{a.get('percentage', 0):.1f}%",
            a.get("started_at_fmt", ""),
            a.get("submitted_at_fmt", ""),
        ]
        ans_map = a.get("answers_by_qid", {})
        for q in questions:
            ans = ans_map.get(q["id"])
            row.append(_render_answer(ans, q))
        w.writerow(row)
    return buf.getvalue()


def attempts_to_xlsx(attempts: list[dict], questions: list[dict], quiz_title: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    header = ["Name", "Email", "Score", "Max", "Percentage", "Started", "Submitted"]
    for q in questions:
        header.append(f"Q{q['position']+1}: {q['text'][:60]}")
    ws.append(header)

    head_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    head_font = Font(color="FFFFFF", bold=True)
    for col in ws[1]:
        col.fill = head_fill
        col.font = head_font
        col.alignment = Alignment(vertical="center", wrap_text=True)

    for a in attempts:
        row = [
            a.get("student_name", ""),
            a.get("student_email", ""),
            a.get("score", 0),
            a.get("max_score", 0),
            a.get("percentage", 0),
            a.get("started_at_fmt", ""),
            a.get("submitted_at_fmt", ""),
        ]
        ans_map = a.get("answers_by_qid", {})
        for q in questions:
            ans = ans_map.get(q["id"])
            row.append(_render_answer(ans, q))
        ws.append(row)

    for i, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(12, max_len + 2), 50)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _render_answer(ans: dict | None, q: dict) -> str:
    if not ans:
        return ""
    try:
        val = json.loads(ans.get("answer") or "null")
    except Exception:
        val = ans.get("answer")
    if val is None:
        return ""
    if q["type"] in ("mcq_single", "mcq_multi", "true_false"):
        options = json.loads(q["options"]) if isinstance(q.get("options"), str) else (q.get("options") or [])
        if isinstance(val, list):
            return ", ".join(options[i] for i in val if 0 <= i < len(options))
        if isinstance(val, int) and 0 <= val < len(options):
            return options[val]
        return str(val)
    if isinstance(val, list):
        return " | ".join(str(v) for v in val)
    return str(val)
